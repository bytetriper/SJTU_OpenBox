from ast import keyword
from cgi import print_directory
from enum import Flag
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import filesearch
import os
def init_filepath(path):
    #path=input("Input a valid path:")
    if os.path.exists(path):
        pass
    else:
        print("[excel search]:Invalid Path")
        exit(-1)
    os.chdir(path)
    files=filesearch.filesearch(path,typename='xlsx',output=True)
    return files
def single_excel_search(key,filename):
    try:
        exc=load_workbook(filename=filename)
    except:
        print("[excel search]:error at opening "+filename+".xlsx")
        return
    sheets=exc.sheetnames
    fg=False
    for sheet in sheets:
        worksheet=exc[sheet]
        i=0
        j=0
        #print(sheet)
        for row in worksheet.iter_rows(values_only=True):
            i+=1
            for value in row:
                j+=1
                if str(key)==str(value):
                    fg=True
                    print("[excel search]:find in [xlsx] file "+filename+"-sheet "+sheet+"-row "+str(i)+":")
                    print(row)
                    break
    if fg==False:
        print("[excel search]:No match in file "+filename)
    return
def excel_search(key,filepath):
    print("[excel search]:searching excels...")
    excelfiles=init_filepath(filepath)
    #print(excelfiles)
    #print("Find excels")
    for file in excelfiles:
        single_excel_search(key,file)
if __name__=='__main__':
    key=input("input a key:")
    path=os.getcwd()
    print(path)
    excel_search(key,path)